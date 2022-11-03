import openpyxl
import xlrd


import openpyxl
import os

from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import config
from message import ShowMsg
temp = 'template.xlsx'
output = "./output.xlsx"
excelfile = './biao.xlsx'


list_read = []
list_new = []
rows = 0
cols = 0

def read_excel():
	"""
	从xlsx文件中读取数据
	"""
	global rows
	global cols
	global list_read
	global list_new

	workbook = openpyxl.load_workbook(filename=temp)
	# print(workbook)
	# 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
	# print(workbook.sheetnames)
	table = workbook.active
	# print(table)

	rows = table.max_row
	cols = table.max_column


	# context = table.cell(2, 3).value
	# ttpye = type(context)
	# print(ttpye)

	# if ttpye == 'str':
	# 	print("type ok")
	# else:
	# 	print("[err]")

	# list1 = [1,2,3]
	# list1.append(4)
	# print("=>被解析表格的总行数 %s, 总列数 %s" % (rows, cols))

	# rows = 7704
	# cols = 8
	dict_conf = config.ReadConfig()
	if dict_conf is None:
		ShowMsg(r'缺少配置文件')
		return
	# print('配置文件读取到数据：行数%d 列数%d' % (dict_conf['row'], dict_conf['col']))
	rows = dict_conf['row'] + 1		# range(a,b)为左闭右开区间，这里为右边最后一位做补偿
	cols = dict_conf['col'] + 1


	list_Fapiao_num = []
	for row in range(2, rows):
		list_per_row = []


		for col in range(1, cols):
			cell_context = table.cell(row, col).value
			# print('%s' % cell_context)
			list_per_row.append(cell_context)

		context = list_per_row[2]
		# print('row%d %s' % (row, context))
		if context is None:
			ShowMsg('行%d，列C 的内容为None\n但是，为便于解析且不影响表格数据，已将其临时修改为空字符，最终输出表格中的对应单元格内容保持不变，依然为None' % row)
			list_per_row[2] = ''
		elif isinstance(list_per_row[2], str):
			list_per_row[2] = list_per_row[2].zfill(8)
		else:
			list_per_row[2] = "%08d" % list_per_row[2]
		# print(context)

		list_read.append(list_per_row)
	# print(list_all)

	# 按第三列排序
	list_read = sorted(list_read, key=(lambda x: x[2]), reverse=False)
	# print("list_Fapiao_num %s " % list_all)



	return

def write_excel():

	global list_read
	global list_new

	# excel文件的写入

	wb = openpyxl.Workbook()
	# 新建了一个工作表，尚未保存

	# print(wb.get_sheet_names())
	# 默认提供Sheet的表，office 2016 默认新建Sheet1

	# 直接赋值可以改工作表的名称
	sheet = wb.active
	sheet.title = 'Sheet1'

	# 可以对工作表Sheet1 Sheet2 建立索引

	# wb.create_sheet('Data', index=1)

	# 这样新建第二个工作表：名称为Data

	# 删除某个工作表
	# remove 按照变量删除
	# wb.remove(sheet)

	# del 按表格名称删除
	# del wb['Data']

	# 新建一个工作表

	# wb = Workbook()
	# grab the active worksheet
	# sheet = wb.active

	# 直接给单元格赋值
	sheet['A1'] = '序号'
	sheet['B1'] = '发票代码'
	sheet['C1'] = '发票号码（定额票也要填）'
	sheet['D1'] = '发票开具单位名称'
	sheet['E1'] = '发票金额'
	sheet['F1'] = '报销部门(项目）'
	sheet['G1'] = '发票登记日期'
	sheet['H1'] = '报销人'


	# 挑选出list_read[2]为空的单元格,并赋值为None
	list_new = list_read
	for row in range(rows - 2):
		if list_read[row][2] == '':
			list_new[row][2] = None

	# append函数
	# 添加一行
	for row in range(rows - 2):
		sheet.append(list_new[row])
	# wb.save('test.xlsx')


	# for i in range(17):
	# 	sheet.cell(row=i+2, column=1, value=i+1)

	# Save the file
	path = output
	# path = filer + "properties.xlsx"
	if os.path.exists(path):
		# print("删除原有的输出文件 %s" % path)
		pass

	# print('输出文件已生成 %s' % path)
	wb.save(path)


	return



def check():

	para = config.ReadConfig()

	workbook = openpyxl.load_workbook(filename=output)
	# print(workbook)
	# 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
	# print(workbook.sheetnames)
	table = workbook.active

	# font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=123456"")  # 设置字体样式

	global list_read
	print("len", len(list_read))

	list_flag = []
	count = para['continuous'] - 1
	for row in range(rows - 2 - count):
		if list_read[row][2] is None or list_read[row + count][2] is None:
			continue
		if int(list_read[row + count][2]) - count == int(list_read[row][2]):
			list_flag.append(row + count)

	# print(list_flag)


	fille = PatternFill('solid', fgColor='FF0000')  # 设置填充颜色为 橙色
	for row in range(rows - 2):
		if row in list_flag:
			table.cell(row + 2, 3).fill = fille

	workbook.save(output)

	return

