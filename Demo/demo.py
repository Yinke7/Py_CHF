import openpyxl
import openpyxl
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块

def Textcolor(file_name,title):
    wk = openpyxl.load_workbook(file_name)  # 加载已经存在的excel
    sheet1 = wk[title]#wk[wk_name[0]]#title名称
    for i in range(10):
        #Color=['c6efce','006100']#绿
        #Color = ['ffc7ce', '9c0006']  #红
        #Color = ['ffeb9c', '9c6500']  # 黄
        Color = ['ffffff', '000000']  # 黑白

        fille = PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
        font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=Color[1])  # 设置字体样式
        sheet1.cell(row=i + 2, column=8, value="").fill = fille  # 序列
        sheet1.cell(row=i + 2, column=8, value="哈哈").font =font # 序列

    wk.save(file_name)  # 保存excel



file_name, title = 's.xlsx', 'Sheet1'
Textcolor(file_name,title)


